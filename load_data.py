import pandas as pd 
import numpy as np
import openpyxl

def load_council_tax_data():
    # Open the Council tax data set
    workbook = openpyxl.load_workbook('Data/ENGLAND_council_tax_band_properties.xlsx')
    
    # read columns that contains the lastest property counts for all bands (2020's) and administrative area names in England.
    council_tax_band = workbook['CTSOP4.0']
    council_tax_band = pd.DataFrame(council_tax_band.values)
    council_tax_band = council_tax_band.drop(labels=range(0,5), axis=0)
    council_tax_band = council_tax_band.drop(columns=range(council_tax_band.columns[6],council_tax_band.columns[31]))
    council_tax_band = council_tax_band.drop(columns=range(council_tax_band.columns[0],council_tax_band.columns[4]))
    council_tax_band = pd.DataFrame(council_tax_band)
    council_tax_band = council_tax_band.rename(columns={4:'area_name', 5:'band', 31:'number_of_properties'})
    
    # Extract the Area Name labels
    area_names = list(pd.unique(council_tax_band[:]['area_name']))
    
    # Extract the properties in each band.
    band_A = council_tax_band[council_tax_band["band"]=='A'].replace('-', 0) # A, Up to and including £40,000, tax charge £1,486.91
    band_B = council_tax_band[council_tax_band["band"]=='B'].replace('-', 0) # B, £40,001 - £52,000, tax charge £1,734.73
    band_C = council_tax_band[council_tax_band["band"]=='C'].replace('-', 0) # C, £52,001 - £68,000, tax charge £1,982.55
    band_D = council_tax_band[council_tax_band["band"]=='D'].replace('-', 0) # D, £68,001 - £88,000, tax charge £2,230.37
    band_E = council_tax_band[council_tax_band["band"]=='E'].replace('-', 0) # E, £88,001 - £120,000, tax charge £2,726.01
    band_F = council_tax_band[council_tax_band["band"]=='F'].replace('-', 0) # F, £120,001 - £160,000, tax charge £3,221.64
    band_G = council_tax_band[council_tax_band["band"]=='G'].replace('-', 0) # G, £160,001 - £320,000, tax charge £3,717.28
    band_H = council_tax_band[council_tax_band["band"]=='H'].replace('-', 0) # H, Over £320,000, tax charge £4,460.74

    # Transform them into nd.array for vector sum purpose.
    band_A = np.array(band_A['number_of_properties']).reshape(-1,1)
    band_B = np.array(band_B['number_of_properties']).reshape(-1,1)
    band_C = np.array(band_C['number_of_properties']).reshape(-1,1)
    band_D = np.array(band_D['number_of_properties']).reshape(-1,1)
    band_E = np.array(band_E['number_of_properties']).reshape(-1,1)
    band_F = np.array(band_F['number_of_properties']).reshape(-1,1)
    band_G = np.array(band_G['number_of_properties']).reshape(-1,1)
    band_H = np.array(band_H['number_of_properties']).reshape(-1,1)
    
    # 8 groups are too many for analysis, thus put them into four groups, catergorized by their valuations.
    # 0 ~ 52k, approximately the houses that are around 50k, mostly small flats or less favoured housing;
    houses_above_0k = pd.DataFrame({'area_names':list(area_names),'property_counts':list((band_A+band_B).reshape(1,-1)[0])})
    # 52 ~ 88k, around 100k, low-average standard;
    houses_above_52k = pd.DataFrame({'area_names':area_names,'property_counts':list((band_C+band_D).reshape(1,-1)[0])})
    # 88k ~ 160k, from 100k to 150k, high-average standard;
    houses_above_88k = pd.DataFrame({'area_names':area_names,'property_counts':list((band_E+band_F).reshape(1,-1)[0])})
    # 160k above, above 150k, more or less luxury houses;
    houses_above_160k = pd.DataFrame({'area_names':area_names,'property_counts':list((band_G+band_H).reshape(1,-1)[0])})
    
    # Return the tables
    return houses_above_0k, houses_above_52k, houses_above_88k, houses_above_160k

def load_homeless_demographic():
    workbook = pd.ExcelFile('Data/England_HomelessData_until2020.xlsx')
    worksheet = workbook.parse(workbook.sheet_names[7])
    hl_demographic = worksheet.drop(labels=range(325,342), axis=0)
    hl_demographic = hl_demographic.drop(columns=hl_demographic.columns[range(11,60)])
    
    hl_demographic.iloc[1,[0]] = 'Code'
    hl_demographic.iloc[1,[1]] = 'Area_name'
    hl_demographic.columns = list(hl_demographic.iloc[1,:])
    hl_demographic = hl_demographic.drop(labels=range(0,16), axis=0)
    hl_demographic.index = range(0,len(hl_demographic.index))
    
    return hl_demographic

def load_time_series_data():
    workbook = openpyxl.load_workbook('Data/all_in_one_Time_sieries_data.xlsx')
    time_series_data = pd.DataFrame(workbook['Usable'].values)
    time_series_data.columns = list(time_series_data.iloc[0])
    time_series_data = time_series_data.drop(labels=0, axis=0)
    afforable_started = time_series_data['Number_of_afforable_houses_started']
    afforable_completed = time_series_data['Number_of_afforable_houses_completed']
    total_started = time_series_data['Total_house_construction_started']
    total_completed = time_series_data['Total_housing_completed']
    timed_waiting_list_size = time_series_data['Size_of_waiting_list']
    timed_median_afforability_ratio = time_series_data['Median_House_price_and_earning_ratio']
    timed_lower_quatile_afforability_ratio = time_series_data['Lower_quatile_House_price_and_earning_ratio']
    
    return time_series_data,afforable_started,afforable_completed,total_started,total_completed,total_started,total_completed,timed_waiting_list_size,timed_median_afforability_ratio,timed_lower_quatile_afforability_ratio


# Everything above is not important, just ignore them!

######################################################################################################################################
##################################################### ALL IN ONE FROM HEER ##########################################################
####################################################################################################################################
def load_categorical_data():
    # How to use?
    # Step 1: from load_data import load_categorical_data
    # Step 2: categorical_data,local_authority_name,total_duty_owed,total_population_in_households,prevention_duty_owed,relief_duty_owed,support_need_homeless,no_longer_homeless,homeless_real_value,categorical_waiting_list_size,social_housing_lettings_2020,band_A_B_properties,band_C_D_properties,band_E_F_properties,band_G_H_properties,count_median_price_houses,count_median_earning_gross,categorical_median_afforability_ratio,count_lower_quatile_price_houses,count_lower_quatile_earning_gross,categorical_lower_quatile_afforability_ratio = load_categorical_data()
    workbook = openpyxl.load_workbook('Data/all_in_one_Categorical_imputed.xlsx')
    categorical_data = pd.DataFrame(workbook['Sheet1'].values)
    categorical_data.columns = list(categorical_data.iloc[0])
    categorical_data = categorical_data.drop(labels=0, axis=0)
    
    local_authority_name = categorical_data['Local_Authority_name']
    total_duty_owed = categorical_data['Total_owed_a_prevention_or_relief_duty'] ##################################### Prevention Duty
    prevention_duty_owed = categorical_data['Threatened_with_homelessness_within_56_days_Prevention_duty_owed'] #### Relief Duty
    relief_duty_owed = categorical_data['Homeless_Relief_duty_owed4']
    total_population_in_households = categorical_data['Number_of_households_in_area4(000s)']
    support_need_homeless = categorical_data['Total_households_with_support_needs']
    
    ## Training target Major
    no_longer_homeless = categorical_data['Total_secured_accommodation'] ########################################## No Longer Homeless
    homeless_real_value = categorical_data['Homeless_(including_intentionally_homeless)']#########################_Homeless Real value
    ## Training target Major
    
    categorical_waiting_list_size = categorical_data['Size_of_social_housing_waiting_list_2020']
    social_housing_lettings_2020 = categorical_data['2020_Total_Lettings']
    band_A_B_properties = categorical_data['A_B_property_counts']
    band_C_D_properties = categorical_data['C_D_property_counts']
    band_E_F_properties = categorical_data['E_F_property_counts']
    band_G_H_properties = categorical_data['G_H_property_counts']
    count_median_price_houses = categorical_data['median_houses_2020']
    count_median_earning_gross = categorical_data['median_earning_2020']
    categorical_median_afforability_ratio = categorical_data['ratio_by_medians_2020']
    count_lower_quatile_price_houses = categorical_data['lower_quatile_houses_2020']
    count_lower_quatile_earning_gross = categorical_data['lower_quatile_earning_2020']
    categorical_lower_quatile_afforability_ratio = categorical_data['ratio_by_lower_quatile_2020']
    
    return categorical_data,local_authority_name,total_duty_owed,total_population_in_households,prevention_duty_owed,relief_duty_owed,support_need_homeless,no_longer_homeless,homeless_real_value,categorical_waiting_list_size,social_housing_lettings_2020,band_A_B_properties,band_C_D_properties,band_E_F_properties,band_G_H_properties,count_median_price_houses,count_median_earning_gross,categorical_median_afforability_ratio,count_lower_quatile_price_houses,count_lower_quatile_earning_gross,categorical_lower_quatile_afforability_ratio

#def load_categorical_more_data():
    # How to use?
    # Step 1: from load_data import load_categorical_more_data
    # Step 2: affordableRent_start,social_housing_start,intermediate_start,total_affordable_start,affordable_complete,social_complete,intermediate_complete,total_affordable_complete = load_categorical_more_data()
    
 #   workbook = openpyxl.load_workbook('Data/all_in_one_Categorical_imputed.xlsx')
   # categorical_data = pd.DataFrame(workbook['Sheet1'].values)
  #  categorical_data.columns = list(categorical_data.iloc[0])
    #categorical_data = categorical_data.drop(labels=0, axis=0)
    
    #a#ffordableRent_start = categorical_data['Affordable_Rent(Starts_on_Site)']
    #social_housing_start = categorical_data['Social_Rent_(Starts_on_Site)']
    #intermediate_start = categorical_data['Intermediate_Rent_(Starts_on_Site)']
    #total_affordable_start = categorical_data['Total_Affordable_(Starts_on_Site)']
    #affordable_complete = categorical_data['Affordable_Rent_(Completions)']
    #social_complete = categorical_data['Social_Rent_(Completions)']
    #intermediate_complete = categorical_data['Intermediate_Rent_(Completions)']
    #total_affordable_complete = categorical_data['Total_Affordable_(Completions)']
    #return affordableRent_start,social_housing_start,intermediate_start,total_affordable_start,affordable_complete,social_complete,intermediate_complete,total_affordable_complete